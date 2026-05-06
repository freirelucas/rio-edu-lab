"""Auditoria do shortlist preliminar de Excels.

Para os top-N candidatos do `excel_catalog.csv` (filtro: span ≥ 5 anos
e max_unique_first_col ≥ 30, ordenado por span desc, views desc), abre
cada arquivo, lê as primeiras linhas e tabula:

  - header real (linhas 0-2)
  - amostra dos rótulos da coluna 0 (linhas 3-12)
  - largura efetiva (cols não-vazias na linha 0 ou 1)
  - veredito de utilidade (USE / NEEDS_CLEANING / SKIP) com justificativa

Outputs:
  - docs/reports/04_shortlist_audit.md
  - data/processed/shortlist_audit.csv
"""

from __future__ import annotations

import csv
import json
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
CATALOG = ROOT / "data" / "processed" / "excel_catalog.csv"
RAW_DIR = ROOT / "data" / "raw" / "excel"
OUT_CSV = ROOT / "data" / "processed" / "shortlist_audit.csv"
REPORT = ROOT / "docs" / "reports" / "04_shortlist_audit.md"

TOP_N = 12
PREVIEW_ROWS = 12
PREVIEW_COLS = 8


def detect_format(path: Path) -> str:
    with path.open("rb") as fh:
        magic = fh.read(8)
    if magic.startswith(b"PK\x03\x04"):
        return "xlsx"
    if magic.startswith(b"\xd0\xcf\x11\xe0"):
        return "xls"
    return "unknown"


def cell_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


# Sheet names that are clearly metadata, not data. No \b at end because
# Portuguese words like "Definições" / "Sumário" have non-ASCII letters
# which Python's \b treats inconsistently around accented chars.
INFO_SHEET_PATTERNS = re.compile(
    r"(metodologia|definiç|notas?\b|glossári|sumári|índice\b|fonte|legenda|capa\b|t[ií]tulo|apresentaç)",
    re.IGNORECASE,
)


def is_info_sheet(name: str) -> bool:
    return bool(INFO_SHEET_PATTERNS.search(name))


def find_header_row(rows: list[list[str]]) -> int:
    """Pick the row with the most non-empty cells among the first 8."""
    best_idx, best_count = 0, -1
    for i, row in enumerate(rows[:8]):
        c = sum(1 for v in row if v)
        if c > best_count:
            best_count = c
            best_idx = i
    return best_idx


def read_xls(path: Path):
    import xlrd

    book = xlrd.open_workbook(str(path), on_demand=True)
    sheets = []
    for idx in range(book.nsheets):
        sh = book.sheet_by_index(idx)
        rows = []
        for r in range(min(sh.nrows, PREVIEW_ROWS)):
            rows.append([cell_str(sh.cell_value(r, c)) for c in range(min(sh.ncols, PREVIEW_COLS))])
        col0 = []
        for r in range(3, min(sh.nrows, 3 + PREVIEW_ROWS)):
            v = cell_str(sh.cell_value(r, 0))
            if v:
                col0.append(v)
        sheets.append({
            "name": sh.name,
            "n_rows": sh.nrows,
            "n_cols": sh.ncols,
            "preview": rows,
            "col0_sample": col0,
        })
        book.unload_sheet(idx)
    book.release_resources()
    return sheets


def read_xlsx(path: Path):
    from openpyxl import load_workbook

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    sheets = []
    try:
        for name in wb.sheetnames:
            ws = wb[name]
            rows = []
            col0 = []
            for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
                if r_idx >= PREVIEW_ROWS + 5:
                    break
                row_strs = [cell_str(v) for v in row[:PREVIEW_COLS]]
                if r_idx < PREVIEW_ROWS:
                    rows.append(row_strs)
                if 3 <= r_idx <= 14 and row and row[0] not in (None, ""):
                    col0.append(cell_str(row[0]))
            sheets.append({
                "name": name,
                "n_rows": ws.max_row or 0,
                "n_cols": ws.max_column or 0,
                "preview": rows,
                "col0_sample": col0[:PREVIEW_ROWS],
            })
    finally:
        wb.close()
    return sheets


def classify_granularity(col0: list[str]) -> str:
    """Heurística mais rigorosa: olha os valores reais, não só o count.

    Volta de regiões administrativas conhecidas do Rio.
    """
    if not col0:
        return "(coluna 0 vazia)"
    sample = " | ".join(col0[:12]).lower()
    # AP/RA usa numeração romana: "I PORTUARIA", "II CENTRO", etc.
    if re.search(r"\b(i+|iv|v|vi+) [a-záéíóúãõçâêô]+", sample):
        return "RA (numeração romana — XX RAs do Rio)"
    if re.search(
        r"\b(centro|copacabana|botafogo|tijuca|jacarepaguá|campo grande|bangu|"
        r"méier|ramos|madureira|santa cruz|portuária|portuaria|barra da tijuca|"
        r"ilha do governador|guaratiba|paquetá|inhaúma|irajá|complexo do alemão|"
        r"vigário geral|anchieta|pavuna|cidade de deus|realengo|penha|ramos)\b",
        sample,
    ):
        return "RA / bairro (nomes geográficos do Rio)"
    if re.search(r"\b\d+ª\s*cre\b|\b[1-9]ª\s*cre\b|\bcre\s*[1-9]+\b", sample):
        return "CRE"
    if any(w in sample for w in ["total", "rio de janeiro"]) and len(col0) <= 5:
        return "(totais / agregado)"
    if all(re.fullmatch(r"\d{4,}", s) for s in col0[:5]):
        return "(códigos numéricos — possível escola/INEP)"
    return "(rótulos heterogêneos — auditar manualmente)"


def pick_data_sheet(sheets: list[dict]) -> dict | None:
    """Skip metadata sheets; pick the first that looks like data."""
    if not sheets:
        return None
    for s in sheets:
        if not is_info_sheet(s["name"]):
            return s
    return sheets[0]


def labels_below_header(sheet: dict, header_row: int) -> list[str]:
    """Return likely label values: pick from cols 0..2, whichever has most
    non-empty text values. Real datasets often put a numeric code in col 0
    and the RA / bairro name in col 1 or 2.
    """
    candidates = []
    for col_idx in range(min(3, max(len(r) for r in sheet["preview"]) if sheet["preview"] else 0)):
        vals = []
        for r in sheet["preview"][header_row + 1 :]:
            if col_idx < len(r) and r[col_idx]:
                vals.append(r[col_idx])
        candidates.append(vals)
    if not candidates:
        return sheet["col0_sample"]
    # Score each by number of unique non-numeric strings
    def score(vals: list[str]) -> int:
        return len({v for v in vals if not re.fullmatch(r"-?\d+(\.\d+)?", v)})
    best = max(candidates, key=score)
    return best if best else sheet["col0_sample"]


def verdict(item: dict, data_sheet: dict | None) -> tuple[str, str, int, list[str]]:
    """Veredito conservador. Returns (verdict, rationale, header_row_idx, col0_sample)."""
    if not data_sheet:
        return "SKIP", "sem sheet de dados (todas parecem metadados)", 0, []
    if data_sheet["n_rows"] < 5:
        return "SKIP", f"sheet primária com {data_sheet['n_rows']} linhas", 0, []
    header_row = find_header_row(data_sheet["preview"])
    h_cells = data_sheet["preview"][header_row] if header_row < len(data_sheet["preview"]) else []
    h_nonempty = sum(1 for c in h_cells if c)
    if h_nonempty < 2:
        return ("SKIP", f"sheet '{data_sheet['name']}' sem cabeçalho identificável", header_row, [])
    col0 = labels_below_header(data_sheet, header_row)
    gran = classify_granularity(col0)
    if "totais" in gran or "vazia" in gran:
        return "SKIP", f"granularidade: {gran}", header_row, col0
    if "auditar" in gran:
        return (
            "NEEDS_CLEANING",
            f"sheet '{data_sheet['name']}' header @ linha {header_row}; "
            f"granularidade não reconhecida ({gran})",
            header_row,
            col0,
        )
    return (
        "USE",
        f"sheet '{data_sheet['name']}' header @ linha {header_row}; "
        f"granularidade: {gran}; {h_nonempty} colunas no header",
        header_row,
        col0,
    )


def md_preview(rows: list[list[str]], max_col_width: int = 24) -> str:
    if not rows:
        return "_(vazio)_"
    n_cols = max(len(r) for r in rows)
    rows = [r + [""] * (n_cols - len(r)) for r in rows]

    def trunc(s: str) -> str:
        s = s.replace("|", "\\|").replace("\n", " ")
        return s if len(s) <= max_col_width else s[: max_col_width - 1] + "…"

    out = ["| " + " | ".join(f"c{i}" for i in range(n_cols)) + " |"]
    out.append("| " + " | ".join(["---"] * n_cols) + " |")
    for r in rows:
        out.append("| " + " | ".join(trunc(c) if c else "·" for c in r) + " |")
    return "\n".join(out)


def main() -> int:
    catalog = list(csv.DictReader(CATALOG.open(encoding="utf-8")))
    excel_items = {r["id"]: r for r in catalog}

    shortlist = sorted(
        [
            r for r in catalog
            if r["years_min"]
            and (int(r["years_max"]) - int(r["years_min"])) >= 5
            and int(r["max_unique_first_col"]) >= 30
        ],
        key=lambda r: (
            int(r["years_max"]) - int(r["years_min"]),
            int(r["num_views"]),
        ),
        reverse=True,
    )[:TOP_N]

    print(f"auditing top {len(shortlist)} of shortlist")

    audit_rows = []
    md_parts: list[str] = []
    md_parts.append("# 04 — Auditoria do shortlist preliminar\n")
    md_parts.append(
        f"Auditoria manual via inspeção dos primeiros {PREVIEW_ROWS} × {PREVIEW_COLS} "
        f"cells de cada sheet primária. Critério para o shortlist: span ≥ 5 anos e ≥ 30 "
        "valores únicos na coluna 0 (Relatório 03). Aqui faço o passo seguinte: olhar o "
        "conteúdo em si para emitir um veredito conservador (`USE` / `NEEDS_CLEANING` / "
        "`SKIP`) sobre se vale levar adiante para o produto HEX-EDU.\n"
    )
    md_parts.append("Tudo que estiver marcado `NEEDS_CLEANING` precisa de "
                    "investigação manual antes de virar input para Theil. `USE` significa "
                    "que estrutura mínima existe — não que os dados estejam validados.\n")

    summary_rows = []

    for r in shortlist:
        item_id = r["id"]
        path = RAW_DIR / f"{item_id}.xlsx"
        if not path.exists():
            print(f"  MISSING {item_id}")
            continue
        fmt = detect_format(path)
        try:
            sheets = read_xls(path) if fmt == "xls" else read_xlsx(path)
        except Exception as e:
            print(f"  PARSE FAIL {item_id}: {e}")
            audit_rows.append({**r, "verdict": "ERROR", "rationale": repr(e), "granularity_real": ""})
            continue

        data_sheet = pick_data_sheet(sheets)
        v, why, header_row, col0 = verdict(r, data_sheet)
        gran_real = classify_granularity(col0)
        audit_rows.append({
            "id": item_id,
            "title": r["title"],
            "format": fmt,
            "n_sheets": int(r["n_sheets"]),
            "years_min": r["years_min"],
            "years_max": r["years_max"],
            "max_unique_first_col": int(r["max_unique_first_col"]),
            "num_views": int(r["num_views"]),
            "data_sheet": data_sheet["name"] if data_sheet else "",
            "header_row": header_row,
            "granularity_real": gran_real,
            "verdict": v,
            "rationale": why,
        })
        summary_rows.append([
            v,
            f"{r['years_min']}–{r['years_max']}",
            r["max_unique_first_col"],
            r["num_views"],
            (r["title"] or "")[:55],
            f"`{item_id[:8]}…`",
        ])

        # Per-file section
        md_parts.append(f"## {v} — {r['title']}\n")
        md_parts.append(
            f"- ID: `{item_id}` · format: `{fmt}` · sheets: {r['n_sheets']} · "
            f"anos: {r['years_min']}–{r['years_max']} · views: {r['num_views']}\n"
        )
        md_parts.append(f"- Veredito: **{v}** — {why}")
        md_parts.append(f"- Granularidade real (col 0 abaixo do header): {gran_real}\n")
        if data_sheet:
            md_parts.append(
                f"### Sheet de dados — `{data_sheet['name']}` "
                f"({data_sheet['n_rows']} × {data_sheet['n_cols']}, header em linha {header_row})\n"
            )
            md_parts.append(md_preview(data_sheet["preview"]))
            md_parts.append("")
            other_names = [s["name"] for s in sheets if s["name"] != data_sheet["name"]]
            if other_names:
                names = ", ".join(f"`{n}`" for n in other_names[:6])
                more = "…" if len(other_names) > 6 else ""
                md_parts.append(f"_Outras sheets: {names}{more}_\n")

    md_parts.insert(2, "## Resumo\n")
    md_parts.insert(3, "Ordenado pela ordem auditada (span desc, views desc).\n")
    md_parts.insert(4, _summary_table(summary_rows))
    md_parts.insert(5, "")

    REPORT.parent.mkdir(parents=True, exist_ok=True)
    REPORT.write_text("\n".join(md_parts), encoding="utf-8")

    OUT_CSV.parent.mkdir(parents=True, exist_ok=True)
    if audit_rows:
        with OUT_CSV.open("w", encoding="utf-8", newline="") as f:
            w = csv.DictWriter(f, fieldnames=list(audit_rows[0].keys()))
            w.writeheader()
            w.writerows(audit_rows)

    print(f"audit complete: {len(audit_rows)} rows")
    print(f"wrote {REPORT.relative_to(ROOT)}")
    print(f"wrote {OUT_CSV.relative_to(ROOT)}")
    return 0


def _summary_table(summary_rows: list[list]) -> str:
    headers = ["Veredito", "Anos", "Unq col 0", "Views", "Título", "ID"]
    out = ["| " + " | ".join(headers) + " |"]
    out.append("| " + " | ".join(["---"] * len(headers)) + " |")
    for row in summary_rows:
        out.append("| " + " | ".join(str(c) for c in row) + " |")
    return "\n".join(out)


if __name__ == "__main__":
    sys.exit(main())
