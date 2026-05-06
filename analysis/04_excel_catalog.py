"""Catálogo empírico dos Excels baixados.

Para cada arquivo em data/raw/excel/:
  - detecta formato real (xls vs xlsx) por magic bytes
  - abre com xlrd ou openpyxl
  - extrai por sheet: dimensões, contagem de valores únicos na 1ª coluna,
    intervalo de anos detectado (4-digit no range [1990, 2030])
  - agrega por arquivo

Outputs:
  - data/processed/excel_catalog.csv      (uma linha por arquivo)
  - data/processed/excel_sheets.csv       (uma linha por sheet)

Uso:
  python analysis/04_excel_catalog.py
"""

from __future__ import annotations

import csv
import json
import re
import sys
import time
from pathlib import Path
from typing import Any, Iterable

ROOT = Path(__file__).resolve().parent.parent
RAW_DIR = ROOT / "data" / "raw" / "excel"
MANIFEST = ROOT / "data" / "manifest.json"
OUT_DIR = ROOT / "data" / "processed"
CATALOG = OUT_DIR / "excel_catalog.csv"
SHEETS = OUT_DIR / "excel_sheets.csv"

YEAR_MIN, YEAR_MAX = 1990, 2030
SCAN_CELLS_PER_SHEET = 5000  # cap para não estourar em planilhas grandes


def detect_format(path: Path) -> str:
    with path.open("rb") as fh:
        magic = fh.read(8)
    if magic.startswith(b"PK\x03\x04"):
        return "xlsx"
    if magic.startswith(b"\xd0\xcf\x11\xe0"):
        return "xls"
    return "unknown"


def years_in_text(text: str) -> set[int]:
    out: set[int] = set()
    for m in re.finditer(r"\b(19\d{2}|20\d{2})\b", text):
        y = int(m.group(0))
        if YEAR_MIN <= y <= YEAR_MAX:
            out.add(y)
    return out


HEADER_ROWS = 4  # numeric values in first N rows are treated as years if in range


def scan_xls(path: Path) -> dict:
    import xlrd

    book = xlrd.open_workbook(str(path), on_demand=True)
    sheets: list[dict] = []
    for idx in range(book.nsheets):
        sh = book.sheet_by_index(idx)
        nrows, ncols = sh.nrows, sh.ncols
        first_col_vals: list[str] = []
        years: set[int] = set()
        cells_scanned = 0
        for r in range(nrows):
            for c in range(ncols):
                if cells_scanned >= SCAN_CELLS_PER_SHEET:
                    break
                v = sh.cell_value(r, c)
                cells_scanned += 1
                if v in (None, ""):
                    continue
                if c == 0:
                    first_col_vals.append(str(v).strip())
                if isinstance(v, (int, float)):
                    iv = int(v)
                    # numeric year only if in header rows (avoids body counts
                    # like 2030 = "número de professores" being read as a year)
                    if r < HEADER_ROWS and YEAR_MIN <= iv <= YEAR_MAX and iv == v:
                        years.add(iv)
                else:
                    years |= years_in_text(str(v))
            if cells_scanned >= SCAN_CELLS_PER_SHEET:
                break
        # First column unique non-empty (skipping header-ish rows: drop top 3 lines)
        body = first_col_vals[3:] if len(first_col_vals) > 3 else first_col_vals
        unique_first = len({v for v in body if v})
        # Header preview: first ~5 non-empty cells of row 0
        header_preview: list[str] = []
        for c in range(min(ncols, 12)):
            try:
                v = sh.cell_value(0, c)
            except Exception:
                v = ""
            if v not in (None, ""):
                header_preview.append(str(v).strip())
        book.unload_sheet(idx)
        sheets.append({
            "sheet_idx": idx,
            "sheet_name": sh.name,
            "n_rows": nrows,
            "n_cols": ncols,
            "n_unique_first_col": unique_first,
            "years_min": min(years) if years else None,
            "years_max": max(years) if years else None,
            "header_preview": " | ".join(header_preview)[:200],
        })
    book.release_resources()
    return {"format": "xls", "sheets": sheets}


def scan_xlsx(path: Path) -> dict:
    from openpyxl import load_workbook

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    sheets: list[dict] = []
    try:
        for idx, name in enumerate(wb.sheetnames):
            ws = wb[name]
            nrows = ws.max_row or 0
            ncols = ws.max_column or 0
            first_col_vals: list[str] = []
            years: set[int] = set()
            cells_scanned = 0
            header_preview: list[str] = []
            for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
                if cells_scanned >= SCAN_CELLS_PER_SHEET:
                    break
                if r_idx == 0:
                    for v in row[:12]:
                        if v not in (None, ""):
                            header_preview.append(str(v).strip())
                if row:
                    first_v = row[0]
                    if first_v not in (None, ""):
                        first_col_vals.append(str(first_v).strip())
                for v in row:
                    cells_scanned += 1
                    if v in (None, ""):
                        continue
                    if isinstance(v, (int, float)):
                        iv = int(v)
                        if r_idx < HEADER_ROWS and YEAR_MIN <= iv <= YEAR_MAX and iv == v:
                            years.add(iv)
                    else:
                        years |= years_in_text(str(v))
                    if cells_scanned >= SCAN_CELLS_PER_SHEET:
                        break
            body = first_col_vals[3:] if len(first_col_vals) > 3 else first_col_vals
            unique_first = len({v for v in body if v})
            sheets.append({
                "sheet_idx": idx,
                "sheet_name": name,
                "n_rows": nrows,
                "n_cols": ncols,
                "n_unique_first_col": unique_first,
                "years_min": min(years) if years else None,
                "years_max": max(years) if years else None,
                "header_preview": " | ".join(header_preview)[:200],
            })
    finally:
        wb.close()
    return {"format": "xlsx", "sheets": sheets}


def aggregate_file(item: dict, parsed: dict, file_bytes: int) -> dict:
    sheets = parsed.get("sheets", [])
    all_years = []
    for s in sheets:
        if s["years_min"] is not None:
            all_years.append(s["years_min"])
        if s["years_max"] is not None:
            all_years.append(s["years_max"])
    return {
        "id": item["id"],
        "title": item.get("title", ""),
        "format": parsed["format"],
        "file_bytes": file_bytes,
        "n_sheets": len(sheets),
        "total_rows": sum(s["n_rows"] for s in sheets),
        "max_cols": max((s["n_cols"] for s in sheets), default=0),
        "max_unique_first_col": max((s["n_unique_first_col"] for s in sheets), default=0),
        "years_min": min(all_years) if all_years else None,
        "years_max": max(all_years) if all_years else None,
        "num_views": item.get("numViews", 0) or 0,
        "tags": "|".join(item.get("tags", []) or []),
    }


def write_csv(path: Path, rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    with path.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader()
        w.writerows(rows)


def main() -> int:
    manifest = json.loads(MANIFEST.read_text(encoding="utf-8"))
    excel_items = {it["id"]: it for it in manifest["items"] if it.get("type") == "Microsoft Excel"}

    files = sorted(RAW_DIR.glob("*.xlsx"))
    print(f"manifest excel items: {len(excel_items)}, files on disk: {len(files)}")

    catalog_rows: list[dict] = []
    sheet_rows: list[dict] = []
    errors: list[tuple[str, str]] = []

    started = time.monotonic()
    for i, path in enumerate(files, 1):
        item_id = path.stem
        item = excel_items.get(item_id)
        if not item:
            errors.append((item_id, "not in manifest as Microsoft Excel"))
            continue

        size = path.stat().st_size
        fmt = detect_format(path)

        try:
            if fmt == "xlsx":
                parsed = scan_xlsx(path)
            elif fmt == "xls":
                parsed = scan_xls(path)
            else:
                errors.append((item_id, f"unknown magic; size={size}"))
                continue
        except Exception as e:
            errors.append((item_id, f"{fmt} parse error: {e!r}"))
            continue

        agg = aggregate_file(item, parsed, size)
        catalog_rows.append(agg)
        for s in parsed["sheets"]:
            sheet_rows.append({"id": item_id, **s})

        if i % 25 == 0 or i == len(files):
            print(f"[{i}/{len(files)}] cataloged  {item_id}  ({fmt}, {len(parsed['sheets'])} sheets)")

    write_csv(CATALOG, catalog_rows)
    write_csv(SHEETS, sheet_rows)

    elapsed = time.monotonic() - started
    print(
        f"\ndone in {elapsed:.1f}s. "
        f"cataloged={len(catalog_rows)}, sheets={len(sheet_rows)}, errors={len(errors)}"
    )
    if errors:
        print("errors:")
        for item_id, msg in errors[:20]:
            print(f"  {item_id}  {msg}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
